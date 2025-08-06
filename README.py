import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import pandas as pd
from time import sleep
import click
import sys
import re

def reduce_upgrade_matrix(input_file, outfile, ver):
    input_file = Path(input_file)
    csvdata = []
    with open(input_file) as file:
        csvreader = csv.reader(file, delimiter=",")
        for row in csvreader:
            csvdata.append(row)
    wb = Workbook()
    ws = wb.active

    ws.title = f"full_matrix_{ver}"

    for row in csvdata:
        for item in row:
            ws.append(item.split("\t"))

    final_column = get_column_letter(ws.max_column)

    frstrow_cell = ws['A1': f"{final_column}1"]

    frstrow_list = []

    for item in frstrow_cell:
        for cell_item in item:
            if cell_item.value is None:
                continue
            else:
                frstrow_list.append(cell_item.value)

    frstcol_cell = ws['A1': f'A{ws.max_row}']

    frstcol_list = []

    for item in frstcol_cell:
        for cell_item in item:
            if cell_item.value is None:
                continue
            else:
                frstcol_list.append(cell_item.value)

    print(f"Target Matrix for : {ver}")

    target_ver = ver.strip().split(".")

    ws_fill_path = wb.create_sheet(title=f"T_values{ver}")

    for idrow, row_val in enumerate(frstrow_list[1:], start=1):
        rval = row_val.strip().split(".")
        if (int(target_ver[0]) == int(rval[0])) | (int(target_ver[0]) - int(rval[0]) == 1) | (
                int(rval[0]) - int(target_ver[0]) == 1):
            ws_fill_path.cell(row=idrow + 1, column=1, value=row_val.strip())
            for idcol, col_val in enumerate(frstcol_list[1:], start=1):
                cval = col_val.strip().split(".")
                if (int(target_ver[0]) == int(cval[0])) | (int(target_ver[0]) - int(cval[0]) == 1) | (
                        int(cval[0]) - int(target_ver[0]) == 1):
                    ws_fill_path.cell(row=1, column=idcol + 1, value=col_val.strip())
                    c = ws.cell(row=idrow + 1, column=idcol + 1)
                    val = c.value
                    if val:
                        ws_fill_path.cell(row=idrow + 1, column=idcol + 1, value=val.strip())

    print(f"x" * 80)
    print(f"Deducted Matrix for :{ver} ")
    print(f"->" * 40)
    print(f"Upgradability from : \n")

    row_count = 0

    for rows in ws_fill_path.iter_rows(min_row=1, max_row=ws_fill_path.max_row, max_col=1):
        for val in rows:
            if val.value is not None:
                print(f"Value is : {val.value}, Cell Coordinate : {val.coordinate}")
                row_count += 1

    print(f"Upgradability to : \n")
    print(f"->" * 40)

    col_count = 0

    for cols in ws_fill_path.iter_cols(max_row=1, min_col=2, max_col=ws_fill_path.max_column):
        for val in cols:
            if val.value is not None:
                print(f"Value is : {val.value}, Cell Coordinate : {val.coordinate}")
                col_count += 1

    print(f"x" * 80)

    target_row_idx = next(i for i in range(2, ws_fill_path.max_row + 1) if ws_fill_path.cell(i, 1).value == ver)
    target_col_idx = next(i for i in range(2, ws_fill_path.max_column + 1) if ws_fill_path.cell(1, i).value == ver)

    for row in ws_fill_path.iter_rows(min_row=2, max_row=ws_fill_path.max_row, min_col=2,
                                      max_col=ws_fill_path.max_column):
        for cell in row:
            if cell.value and (
                    cell.row != target_row_idx or cell.column != target_col_idx) and cell.value.strip().upper() == "F":
                cell.value = ""

    # Remove empty rows and columns

    max_row = ws_fill_path.max_row
    max_col = ws_fill_path.max_column

    rows_to_delete = []

    for row in ws_fill_path.iter_rows(min_row=2, max_row=max_row):
        if all(cell.value is None for cell in row):
            rows_to_delete.append(row[0].row)

    for row in reversed(rows_to_delete):
        ws_fill_path.delete_rows(row)

    cols_to_delete = []

    for col in ws_fill_path.iter_cols(min_col=2, max_col=max_col):
        if all(cell.value is None for cell in col):
            cols_to_delete.append(col[0].column)

    for col in reversed(cols_to_delete):
        ws_fill_path.delete_cols(col)

    print(f"x" * 80)
    print(f"Count of Non-Empty Columns: {col_count}\n")
    print(f"Count of Non-Empty Rows: {row_count}\n")
    print(f"x" * 80)

    wb.active = ws_fill_path
    wb.save(outfile)
    wb.close()


def pd_analyse_reduce(xlsx_path, wsheet_name, ver, out_file):
    df = pd.read_excel(xlsx_path, sheet_name=wsheet_name)

    if ver not in df.columns:
        print(f"{ver} not found in the provided Matrix\n")
        sys.exit(1)

    original_columns = df.columns.tolist()  # Save the original order of columns

    df.dropna(how="all", inplace=True)
    df.dropna(axis=1, how="all", inplace=True)

    df1 = df.sort_index(axis=1)

    r, c = df1.shape

    df1.sort_values(by=df1.columns[c - 1], inplace=True)

    col_list = df1.columns.to_list()

    if ver in col_list:
        df1 = df1[df1[ver].notna()]
    else:
        print(f"{ver} not found in {xlsx_path}")

    print(df1.shape)

    df1.reset_index(drop=True, inplace=True)

    unnamed_cols = [col for col in df1.columns if 'Unnamed' in col]

    if unnamed_cols:
        for col in unnamed_cols:
            cols = [col] + [c for c in df1.columns if c != col]
            df1 = df1[cols]

    target_row = df1.index[df1[df1.columns[0]] == ver][0]
    target_col = df1.columns.get_loc(ver)

    columns_with_T = []

    for value in df1.iloc[target_row]:
        if value == 'T':
            columns_with_T.append(df1.columns[df1.iloc[target_row] == 'T'].tolist())

    columns_with_T = [item for sublist in columns_with_T for item in sublist]

    if ver not in columns_with_T:
        columns_with_T.append(ver)

    columns_to_keep = unnamed_cols + columns_with_T
    df1 = df1[columns_to_keep]

    # Ensure only the columns that are still present in df1 are reordered
    valid_columns = [col for col in original_columns if col in df1.columns]
    df1 = df1[valid_columns]

    with pd.ExcelWriter(xlsx_path, mode="a", engine="openpyxl") as file:
        df1.to_excel(file, sheet_name=f"filtered_path{ver}", index=False)


def create_tracker_sheet(xlsx_path, wsheet_name, target_version):
    df = pd.read_excel(xlsx_path, sheet_name=wsheet_name)

    def version_key(version):
        parts = re.split(r'[.-]', version)
        keys = []
        for part in parts:
            if part.isdigit():
                keys.append(int(part))
            else:
                keys.append(part)
        return keys

    original_columns = df.columns.tolist()  # Save the original order of columns

    df.sort_values(by=df.columns[0], key=lambda col: [version_key(val) for val in col], inplace=True)

    latest_versions = []
    seen_trains = set()

    for version in df[df.columns[0]]:
        train = '.'.join(version.split('.')[:2])
        if train not in seen_trains:
            latest_versions.append(version)
            seen_trains.add(train)
        elif 'beta' in version:
            pass
        elif 'p' in version:
            latest_versions[-1] = version
        else:
            seen_trains.remove(train)
            latest_versions.append(version)
            seen_trains.add(train)

    df = df[df[df.columns[0]].isin(latest_versions)]

    # Ensure only the columns that are still present in df are reordered
    valid_columns = [col for col in original_columns if col in df.columns]
    df = df[valid_columns]

    # Identify the target version row
    target_version_row = df.index[df[df.columns[0]] == target_version].tolist()
    target_version_row = target_version_row[0] if target_version_row else None

    if target_version_row is not None:
        for col in df.columns[2:]:
            df.loc[df.index != target_version_row, col] = df.loc[df.index != target_version_row, col].replace('T', '')

    # Clear cells with 'F' values
    df.replace('F', '', inplace=True)

    with pd.ExcelWriter(xlsx_path, mode="a", engine="openpyxl") as file:
        df.to_excel(file, sheet_name="tracker", index=False)


@click.command()
@click.option("-v", "--ver", required=True, help="Please input the version for which matrix reduction needs to be done")
@click.option("--csv_file_path", required=True, help="Full path for the matrix CSV file")
def analyse_matrix(ver, csv_file_path):
    csv_file = Path(csv_file_path)

    if not csv_file.is_file():
        print("Please provide a valid file path")
        sys.exit(1)

    mtrx_out_file = Path.home() / f"Downloads/results_{ver}_values.xlsx"
    mtrx_final = Path.home() / f"Downloads/reduced{ver}_output.xlsx"
    mtrx_sheet = f"T_values{ver}"

    reduce_upgrade_matrix(csv_file, mtrx_out_file, ver)
    sleep(60)
    pd_analyse_reduce(mtrx_out_file, mtrx_sheet, ver, mtrx_final)
    create_tracker_sheet(mtrx_out_file, f"filtered_path{ver}", ver)

    print(f"Please review output file for Upgrade Matrix Reduction: {mtrx_out_file}\n")
    print(f"Tracker sheet created successfully!\n")


if __name__ == '__main__':
    analyse_matrix()
